"""Excel export of extraction results.

The workbook is built entirely in memory and streamed to the client —
nothing touches disk, keeping the app stateless.
"""

import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from app.dataset import DATASET_COLUMNS, DATASET_LABELS
from app.schemas import FIELD_KEYS, FIELD_LABELS, LabelResult

_HEADER_FILL = PatternFill("solid", fgColor="0B5394")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_VERDICT_STYLES = {
    "pass": (PatternFill("solid", fgColor="C6EFCE"), Font(bold=True, color="1A7F37")),
    "fail": (PatternFill("solid", fgColor="FFC7CE"), Font(bold=True, color="C62828")),
    "warning": (PatternFill("solid", fgColor="FFEB9C"), Font(bold=True, color="9A6700")),
    "error": (PatternFill("solid", fgColor="D9D9D9"), Font(bold=True, color="57606A")),
}

_VERDICT_TEXT = {
    "pass": "PASS",
    "fail": "FAIL",
    "warning": "NEEDS REVIEW",
    "error": "COULD NOT CHECK",
}

_MAX_COLUMN_WIDTH = 60


def _notes_for(result: LabelResult) -> str:
    if result.error:
        return result.error
    notes = [f"{f.label}: {f.notes}" for f in result.fields if f.notes]
    if result.warning_statement:
        notes.extend(result.warning_statement.notes)
    if result.match and result.match.note:
        notes.append(result.match.note)
    notes.extend(result.notes)
    return " | ".join(notes)


CATALOG_SHEET = "Product Catalog"

_LINK_FONT = Font(color="0B5394", underline="single")


def _write_catalog_sheet(ws, products: list[dict]) -> None:
    ws.append([DATASET_LABELS[key] for key in DATASET_COLUMNS])
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for product in products:
        ws.append([product.get(key, "") for key in DATASET_COLUMNS])

    widths = [14, 30, 36, 28, 16, 56, 18]
    for column_index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column_index)].width = width


def build_dataset_workbook(products: list[dict]) -> bytes:
    """The product catalog as a styled .xlsx — reviewers can analyze or
    modify it, add rows, and re-upload it as the matching dataset."""
    wb = Workbook()
    ws = wb.active
    ws.title = CATALOG_SHEET
    _write_catalog_sheet(ws, products)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


IMAGES_SHEET = "Label Images"
_THUMB_MAX = (320, 260)


def _add_images_sheet(wb: Workbook, images: dict[str, bytes]) -> dict[str, int]:
    """Embed a thumbnail per label on their own sheet; returns filename →
    anchor row so results rows can hyperlink to their image."""
    ws = wb.create_sheet(IMAGES_SHEET)
    ws.append(["Label File", "Image"])
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 46

    anchors: dict[str, int] = {}
    row = 2
    for filename, data in images.items():
        try:
            pil = PILImage.open(io.BytesIO(data))
            pil.thumbnail(_THUMB_MAX)
            if pil.mode != "RGB":
                pil = pil.convert("RGB")
            buffer = io.BytesIO()
            pil.save(buffer, format="JPEG", quality=80)
            buffer.seek(0)
        except OSError:
            continue  # undecodable upload — no thumbnail, no link
        ws.cell(row=row, column=1, value=filename).alignment = Alignment(vertical="top")
        image = XLImage(buffer)
        ws.add_image(image, f"B{row}")
        ws.row_dimensions[row].height = pil.height * 0.75 + 8
        anchors[filename] = row
        row += 1
    return anchors


def build_workbook(
    results: list[LabelResult],
    catalog: list[dict] | None = None,
    total_seconds: float | None = None,
    images: dict[str, bytes] | None = None,
    client_seconds: float | None = None,
) -> bytes:
    """Return the bytes of a styled .xlsx: the product catalog as the first
    sheet (when matching was on), then one results row per label, with each
    matched product hyperlinked to its catalog row."""
    catalog = catalog or []

    wb = Workbook()
    if catalog:
        catalog_ws = wb.active
        catalog_ws.title = CATALOG_SHEET
        _write_catalog_sheet(catalog_ws, catalog)
        ws = wb.create_sheet("Label Extraction Results")
    else:
        # Matching was off — no catalog sheet, results only.
        ws = wb.active
        ws.title = "Label Extraction Results"

    # Catalog row each product_id lives on (row 1 is the header).
    catalog_rows = {
        str(product.get("product_id", "")): row
        for row, product in enumerate(catalog, start=2)
    }

    # Match columns only appear when catalog matching ran.
    include_match = bool(catalog)
    headers = (
        ["Label File"]
        + [FIELD_LABELS[key] for key in FIELD_KEYS]
        + ["Government Warning"]
        + (["Matched Product", "Match Score", "Judge Confidence"] if include_match else [])
        + ["Verdict", "Notes", "Processing Time (s)"]
    )
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    verdict_column = len(headers) - 2  # 1-based index of the Verdict column

    # Thumbnails of the checked labels live on their own sheet; each results
    # row's "Label File" cell links to its image.
    image_anchors = _add_images_sheet(wb, images) if images else {}

    for result in results:
        field_map = {f.field: f for f in result.fields}

        def display(key):
            f = field_map.get(key)
            if f is None:
                return ""
            if f.value:
                return f.value
            # Missing but not required (e.g. country of origin on a
            # domestic product) reads as N/A, not as a failure.
            return "N/A" if f.status == "pass" else "NOT FOUND"

        warning = result.warning_statement
        match = result.match
        if warning is None:
            warning_text = ""
        elif warning.extracted_text:
            warning_text = warning.extracted_text
        else:
            warning_text = "NOT FOUND"

        match_cells = []
        if include_match:
            if match is None:
                matched_product = ""
            elif match.matched and match.product:
                matched_product = (
                    f"{match.product.get('product_id', '')} "
                    f"{match.product.get('brand_name', '')}".strip()
                )
            else:
                matched_product = "NO MATCH"
            match_cells = [
                matched_product,
                round(match.score, 3) if match and match.score is not None else "",
                round(match.judge_confidence, 2)
                if match and match.judge_confidence is not None
                else "",
            ]

        row = (
            [result.filename]
            + [display(key) for key in FIELD_KEYS]
            + [warning_text]
            + match_cells
            + [
                _VERDICT_TEXT[result.verdict],
                _notes_for(result),
                round(result.processing_seconds, 1)
                if result.processing_seconds is not None
                else "",
            ]
        )
        if result.error:
            for key_index in range(2, 2 + len(FIELD_KEYS)):
                row[key_index - 1] = ""
        ws.append(row)

        fill, font = _VERDICT_STYLES[result.verdict]
        cell = ws.cell(row=ws.max_row, column=verdict_column)
        cell.fill = fill
        cell.font = font

        # The warning column carries the verbatim transcription; flag only
        # problems (no fill when it passes) and wrap the long text.
        if warning is not None:
            warning_cell = ws.cell(row=ws.max_row, column=len(FIELD_KEYS) + 2)
            if warning.status != "pass":
                warning_cell.fill = _VERDICT_STYLES[warning.status][0]
            warning_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Link the filename to its thumbnail on the images sheet.
        anchor_row = image_anchors.get(result.filename)
        if anchor_row:
            file_cell = ws.cell(row=ws.max_row, column=1)
            file_cell.hyperlink = f"#'{IMAGES_SHEET}'!A{anchor_row}"
            file_cell.font = _LINK_FONT

        # Link the matched product back to its row on the catalog sheet.
        if match is not None and match.matched and match.product:
            catalog_row = catalog_rows.get(str(match.product.get("product_id", "")))
            if catalog_row:
                product_cell = ws.cell(row=ws.max_row, column=len(FIELD_KEYS) + 3)
                product_cell.hyperlink = f"#'{CATALOG_SHEET}'!A{catalog_row}"
                product_cell.font = _LINK_FONT

        # Color failing/review field cells (catalog mismatches and intrinsic
        # failures alike) so problems are scannable column by column.
        for offset, key in enumerate(FIELD_KEYS):
            field = field_map.get(key)
            if field is not None and field.status != "pass":
                field_fill, _ = _VERDICT_STYLES[field.status]
                ws.cell(row=ws.max_row, column=2 + offset).fill = field_fill

    # Per-label times overlap (labels are processed concurrently), so they
    # don't sum to the wall clock — state the totals explicitly, and
    # distinguish processing from the upload-inclusive end-to-end time.
    footer_rows = []
    if total_seconds is not None:
        footer_rows.append((
            "Server processing total — labels run concurrently, so per-label times overlap:",
            round(total_seconds, 1),
        ))
    if client_seconds is not None:
        footer_rows.append((
            "End-to-end request time (incl. uploading the images to the server):",
            round(client_seconds, 1),
        ))
    for text, value in footer_rows:
        footer = [""] * len(headers)
        footer[-2] = text
        footer[-1] = value
        ws.append(footer)
        for cell in ws[ws.max_row]:
            cell.font = Font(italic=True, color="57606A")

    # Auto-width columns from their longest value, capped to stay readable.
    for column_index in range(1, len(headers) + 1):
        letter = get_column_letter(column_index)
        longest = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in ws[letter]
        )
        ws.column_dimensions[letter].width = min(longest + 3, _MAX_COLUMN_WIDTH)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
