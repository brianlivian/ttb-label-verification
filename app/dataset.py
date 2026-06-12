"""The base product catalog and dataset-file parsing.

The catalog is a fictitious dataset of popular liquors — real-sounding
brands with made-up specification values — that stands in for the
application-of-record system (COLAs Online) this prototype would integrate
with in production. Reviewers can download it as Excel, analyze or modify
it, add rows, and re-upload; an uploaded dataset replaces the built-in one
for that request.
"""

import csv
import io
import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

DATASET_COLUMNS = (
    "product_id",
    "brand_name",
    "class_type",
    "alcohol_content",
    "net_contents",
    "bottler_address",
    "country_of_origin",
)

DATASET_LABELS = {
    "product_id": "Product ID",
    "brand_name": "Brand Name",
    "class_type": "Class / Type",
    "alcohol_content": "Alcohol Content",
    "net_contents": "Net Contents",
    "bottler_address": "Bottler Name & Address",
    "country_of_origin": "Country of Origin",
}

# Fictitious specs for 42 products: 22 popular liquors plus the 20
# products behind the generated evaluation labels. IDs are stable so re-uploaded,
# modified datasets can keep referring to the same products.
BASE_PRODUCTS: list[dict[str, str]] = [
    {"product_id": "TTB-001", "brand_name": "Jack Daniel's Old No. 7", "class_type": "Tennessee Whiskey", "alcohol_content": "40% Alc./Vol.", "net_contents": "700 mL", "bottler_address": "Distilled and Bottled by Jack Daniel Distillery, Lynchburg, TN", "country_of_origin": ""},
    {"product_id": "TTB-002", "brand_name": "Johnnie Walker Black Label", "class_type": "Blended Scotch Whisky", "alcohol_content": "40% Alc./Vol. (80 Proof)", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled in Scotland by John Walker and Sons, Kilmarnock, Scotland", "country_of_origin": "Scotland"},
    {"product_id": "TTB-003", "brand_name": "Jameson", "class_type": "Irish Whiskey", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Distilled by John Jameson and Son, Midleton, County Cork, Ireland", "country_of_origin": "Ireland"},
    {"product_id": "TTB-004", "brand_name": "Jim Beam", "class_type": "Kentucky Straight Bourbon Whiskey", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled by James B. Beam Distilling Co., Clermont, KY", "country_of_origin": ""},
    {"product_id": "TTB-005", "brand_name": "Maker's Mark", "class_type": "Kentucky Straight Bourbon Whisky", "alcohol_content": "45% Alc./Vol. (90 Proof)", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled by Maker's Mark Distillery, Loretto, KY", "country_of_origin": ""},
    {"product_id": "TTB-006", "brand_name": "Old Rip Van Winkle", "class_type": "Kentucky Straight Bourbon Whiskey", "alcohol_content": "45.2% Alc./Vol. (90.4 Proof)", "net_contents": "750 mL", "bottler_address": "Bottled by Old Rip Van Winkle Distillery, Frankfort, KY", "country_of_origin": ""},
    {"product_id": "TTB-007", "brand_name": "Crown Royal", "class_type": "Canadian Whisky", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Blended and Bottled by The Crown Royal Distilling Co., Gimli, Manitoba", "country_of_origin": "Canada"},
    {"product_id": "TTB-008", "brand_name": "The Macallan 12", "class_type": "Single Malt Scotch Whisky", "alcohol_content": "43% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled by The Macallan Distillers Ltd., Craigellachie, Scotland", "country_of_origin": "Scotland"},
    {"product_id": "TTB-009", "brand_name": "Smirnoff No. 21", "class_type": "Vodka", "alcohol_content": "37.5% Vol.", "net_contents": "70 cl", "bottler_address": "Produced and Bottled by The Pierre Smirnoff Co., United Kingdom", "country_of_origin": "United Kingdom"},
    {"product_id": "TTB-010", "brand_name": "Absolut", "class_type": "Vodka", "alcohol_content": "40% Alc./Vol. (80 Proof)", "net_contents": "750 mL", "bottler_address": "Produced and Bottled in Ahus, Sweden", "country_of_origin": "Sweden"},
    {"product_id": "TTB-011", "brand_name": "Grey Goose", "class_type": "Vodka", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Bottled by Grey Goose, Cognac, France", "country_of_origin": "France"},
    {"product_id": "TTB-012", "brand_name": "Tito's Handmade", "class_type": "Vodka", "alcohol_content": "40% Alc./Vol. (80 Proof)", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled by Fifth Generation Inc., Austin, TX", "country_of_origin": ""},
    {"product_id": "TTB-013", "brand_name": "Russian Standard Original", "class_type": "Vodka", "alcohol_content": "40% Alc. by Vol. (80 Proof)", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled by Russian Standard, St. Petersburg, Russia", "country_of_origin": "Russia"},
    {"product_id": "TTB-014", "brand_name": "Bacardí Superior", "class_type": "Puerto Rican Rum", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Bottled by Bacardi Corporation, San Juan, Puerto Rico", "country_of_origin": ""},
    {"product_id": "TTB-015", "brand_name": "Bacardí Gran Reserva Diez", "class_type": "Extra Rare Gold Rum", "alcohol_content": "40% Alc./Vol.", "net_contents": "70 cl", "bottler_address": "Bottled by Bacardi Corporation, San Juan, Puerto Rico", "country_of_origin": ""},
    {"product_id": "TTB-016", "brand_name": "Captain Morgan Original Spiced", "class_type": "Rum with Spice and Natural Flavors", "alcohol_content": "35% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Bottled by Captain Morgan Rum Co., New York, NY", "country_of_origin": ""},
    {"product_id": "TTB-017", "brand_name": "Casamigos Blanco", "class_type": "Tequila Blanco 100% Agave", "alcohol_content": "40% Alc./Vol. (80 Proof)", "net_contents": "750 mL", "bottler_address": "Produced and Bottled by Casamigos Tequila, Jalisco, Mexico", "country_of_origin": "Mexico"},
    {"product_id": "TTB-018", "brand_name": "Patrón Silver", "class_type": "Tequila 100% de Agave", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Produced by Hacienda Patron, Atotonilco el Alto, Jalisco, Mexico", "country_of_origin": "Mexico"},
    {"product_id": "TTB-019", "brand_name": "Conmemorativo", "class_type": "Tequila Añejo", "alcohol_content": "40% Alc./Vol.", "net_contents": "700 mL", "bottler_address": "Produced by Tequila San Matias de Jalisco, Mexico", "country_of_origin": "Mexico"},
    {"product_id": "TTB-020", "brand_name": "Tanqueray London Dry", "class_type": "Distilled Gin", "alcohol_content": "47.3% Alc./Vol. (94.6 Proof)", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled in Cameronbridge, Scotland", "country_of_origin": "Scotland"},
    {"product_id": "TTB-021", "brand_name": "Hendrick's", "class_type": "Gin", "alcohol_content": "44% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled by The Girvan Distillery, Girvan, Scotland", "country_of_origin": "Scotland"},
    {"product_id": "TTB-022", "brand_name": "Jägermeister", "class_type": "Herbal Liqueur (Kräuterlikör)", "alcohol_content": "35% Vol.", "net_contents": "0.04 L", "bottler_address": "Produced and Bottled by Mast-Jagermeister SE, Wolfenbuttel, Germany", "country_of_origin": "Germany"},

    # Fictitious products behind the AI-generated test labels in
    # test-data/generated/ (see its manifest.json for planted defects).
    {"product_id": "TTB-101", "brand_name": "Copper Canyon", "class_type": "Straight Rye Whiskey", "alcohol_content": "45% Alc./Vol. (90 Proof)", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled by Copper Canyon Distilling Co., Boulder, CO", "country_of_origin": ""},
    {"product_id": "TTB-102", "brand_name": "Silver Pine", "class_type": "Vodka", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Distilled by Silver Pine Spirits, Boise, ID", "country_of_origin": ""},
    {"product_id": "TTB-103", "brand_name": "Harbor Light", "class_type": "London Dry Gin", "alcohol_content": "47% Alc./Vol. (94 Proof)", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled by Harbor Light Gin Works, Portland, ME", "country_of_origin": ""},
    {"product_id": "TTB-104", "brand_name": "Casa Dorada", "class_type": "Tequila Reposado 100% Agave", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Produced and Bottled by Destileria Casa Dorada, Jalisco, Mexico", "country_of_origin": "Mexico"},
    {"product_id": "TTB-105", "brand_name": "Blackwater Bay", "class_type": "Caribbean Spiced Rum", "alcohol_content": "35% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Bottled by Blackwater Bay Rum Co., Charleston, SC", "country_of_origin": ""},
    {"product_id": "TTB-106", "brand_name": "Old Cooper's", "class_type": "Kentucky Straight Bourbon Whiskey", "alcohol_content": "50% Alc./Vol. (100 Proof)", "net_contents": "375 mL", "bottler_address": "Distilled and Bottled by Old Cooper's Distillery, Bardstown, KY", "country_of_origin": ""},
    {"product_id": "TTB-107", "brand_name": "Glen Marrow", "class_type": "Single Malt Scotch Whisky", "alcohol_content": "43% Alc./Vol.", "net_contents": "700 mL", "bottler_address": "Distilled and Bottled in Scotland by Glen Marrow Distillers, Speyside, Scotland", "country_of_origin": "Scotland"},
    {"product_id": "TTB-108", "brand_name": "Frost Peak", "class_type": "American Dry Gin", "alcohol_content": "44% Alc./Vol.", "net_contents": "1 L", "bottler_address": "Distilled by Frost Peak Botanicals, Bozeman, MT", "country_of_origin": ""},
    {"product_id": "TTB-109", "brand_name": "Rio Verde", "class_type": "Mezcal Artesanal", "alcohol_content": "46% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Produced by Mezcaleria Rio Verde, Oaxaca, Mexico", "country_of_origin": "Mexico"},
    {"product_id": "TTB-110", "brand_name": "Golden Prairie", "class_type": "Lager Beer", "alcohol_content": "4.8% Alc./Vol.", "net_contents": "12 FL OZ", "bottler_address": "Brewed and Bottled by Golden Prairie Brewing Co., Omaha, NE", "country_of_origin": ""},
    {"product_id": "TTB-111", "brand_name": "Willow Creek", "class_type": "Red Table Wine", "alcohol_content": "12.5% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Produced and Bottled by Willow Creek Winery, Napa, CA", "country_of_origin": ""},
    {"product_id": "TTB-112", "brand_name": "North Fork", "class_type": "Apple Brandy", "alcohol_content": "42% Alc./Vol. (84 Proof)", "net_contents": "750 mL", "bottler_address": "Distilled by North Fork Orchards, Hudson, NY", "country_of_origin": ""},
    {"product_id": "TTB-113", "brand_name": "Iron Gate", "class_type": "Vodka", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Distilled by Iron Gate Distillery, Detroit, MI", "country_of_origin": ""},
    {"product_id": "TTB-114", "brand_name": "Cedar Hollow", "class_type": "Tennessee Whiskey", "alcohol_content": "43% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled by Cedar Hollow Distillery, Lynchburg, TN", "country_of_origin": ""},
    {"product_id": "TTB-115", "brand_name": "Santa Lucia", "class_type": "Tequila Blanco", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Produced and Bottled by Tequilera Santa Lucia, Jalisco, Mexico", "country_of_origin": "Mexico"},
    {"product_id": "TTB-116", "brand_name": "Bay & Anchor", "class_type": "Navy Strength Gin", "alcohol_content": "57% Alc./Vol. (114 Proof)", "net_contents": "750 mL", "bottler_address": "Distilled by Bay & Anchor Spirits, Annapolis, MD", "country_of_origin": ""},
    {"product_id": "TTB-117", "brand_name": "Kings Hollow", "class_type": "Straight Bourbon Whiskey", "alcohol_content": "45% Alc./Vol. (90 Proof)", "net_contents": "750 mL", "bottler_address": "Distilled and Bottled by Kings Hollow Distilling, Frankfort, KY", "country_of_origin": ""},
    {"product_id": "TTB-118", "brand_name": "White Falcon", "class_type": "Vodka", "alcohol_content": "40% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Distilled by White Falcon Spirits, Fargo, ND", "country_of_origin": ""},
    {"product_id": "TTB-119", "brand_name": "Stone Mill", "class_type": "Straight Bourbon Whiskey", "alcohol_content": "45% Alc./Vol.", "net_contents": "750 mL", "bottler_address": "Distilled by Stone Mill Distillery, Lancaster, PA", "country_of_origin": ""},
    {"product_id": "TTB-120", "brand_name": "Kiyomi", "class_type": "Junmai Sake", "alcohol_content": "15% Alc./Vol.", "net_contents": "720 mL", "bottler_address": "Brewed and Bottled by Kiyomi Shuzo, Kyoto, Japan", "country_of_origin": ""},
]


class DatasetFileError(Exception):
    """Raised with a user-facing message when an uploaded dataset is unusable."""


# Accepted spellings for each column header, lowercase, punctuation stripped.
_HEADER_ALIASES = {
    "product_id": {"product id", "id", "productid", "sku"},
    "brand_name": {"brand", "brand name", "brandname"},
    "class_type": {
        "class", "type", "class type", "classtype", "class and type",
        "class and type designation", "designation", "class type designation",
    },
    "alcohol_content": {
        "alcohol", "alcohol content", "abv", "alc", "alcohol percent",
        "alcohol percentage", "alcohol abv",
    },
    "net_contents": {
        "net contents", "net content", "netcontents", "contents", "volume", "size",
    },
    "bottler_address": {
        "bottler", "bottler name address", "bottler name and address",
        "name and address", "bottler address", "producer", "bottler producer",
    },
    "country_of_origin": {
        "country", "country of origin", "origin",
    },
}


def _norm_header(text: str) -> str:
    text = re.sub(r"[^a-z0-9 ]", " ", str(text).lower())
    return re.sub(r"\s+", " ", text).strip()


def _rows_from_csv(data: bytes) -> list[list[str]]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    return [row for row in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(data: bytes) -> list[list[str]]:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise DatasetFileError(
            "The Excel file could not be opened. Please save it as .xlsx and try again."
        ) from exc
    ws = wb.active
    return [["" if v is None else str(v) for v in row] for row in ws.iter_rows(values_only=True)]


def parse_dataset_file(data: bytes, filename: str) -> list[dict[str, str]]:
    """Parse an uploaded CSV/XLSX catalog into product rows."""
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        raw_rows = _rows_from_csv(data)
    elif suffix == ".xlsx":
        raw_rows = _rows_from_xlsx(data)
    else:
        raise DatasetFileError(
            f"'{filename}' is not a supported dataset file. Please upload "
            "a .csv or .xlsx file."
        )

    rows = [row for row in raw_rows if any(str(cell).strip() for cell in row)]
    if not rows:
        raise DatasetFileError("The dataset file is empty.")

    columns: dict[int, str] = {}
    for index, cell in enumerate(rows[0]):
        normalized = _norm_header(cell)
        for key, aliases in _HEADER_ALIASES.items():
            if normalized in aliases:
                columns[index] = key
                break

    if "brand_name" not in columns.values():
        raise DatasetFileError(
            "No 'brand name' column found in the dataset file. Use headers "
            "like: product id, brand name, class/type, alcohol content, "
            "net contents (download the base dataset for the exact format)."
        )

    products: list[dict[str, str]] = []
    for line_number, row in enumerate(rows[1:], start=2):
        product = {key: "" for key in DATASET_COLUMNS}
        for index, key in columns.items():
            if index < len(row):
                product[key] = re.sub(r"\s+", " ", str(row[index])).strip()
        if not product["brand_name"]:
            continue  # skip blank/partial rows
        if not product["product_id"]:
            product["product_id"] = f"ROW-{line_number}"
        products.append(product)

    if not products:
        raise DatasetFileError("No usable product rows found in the dataset file.")
    return products


def expected_values(product: dict[str, str]) -> dict[str, str]:
    """The comparison fields of a catalog row (drops the ID and blanks)."""
    return {
        key: product[key]
        for key in DATASET_COLUMNS
        if key != "product_id" and product.get(key)
    }
