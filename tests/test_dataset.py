"""Unit tests for the base catalog and dataset-file parsing."""

import io

import pytest
from openpyxl import load_workbook

from app.dataset import (
    BASE_PRODUCTS,
    DatasetFileError,
    expected_values,
    parse_dataset_file,
)
from app.export import build_dataset_workbook


def test_base_dataset_has_at_least_20_products():
    assert len(BASE_PRODUCTS) >= 20
    ids = [p["product_id"] for p in BASE_PRODUCTS]
    assert len(ids) == len(set(ids))
    for product in BASE_PRODUCTS:
        assert product["brand_name"]
        assert product["class_type"]
        assert product["alcohol_content"]
        assert product["net_contents"]


def test_dataset_workbook_round_trip():
    """Download → modify → add a row → re-upload must work unchanged."""
    data = build_dataset_workbook(BASE_PRODUCTS)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.max_row == len(BASE_PRODUCTS) + 1

    # Modify a value and append a new product, as a reviewer would.
    ws["D2"] = "50% Alc./Vol."
    ws.append(["TTB-099", "Stone's Throw", "Small Batch Gin", "45% ABV", "750 mL"])
    out = io.BytesIO()
    wb.save(out)

    products = parse_dataset_file(out.getvalue(), "modified.xlsx")
    assert len(products) == len(BASE_PRODUCTS) + 1
    assert products[0]["alcohol_content"] == "50% Alc./Vol."
    assert products[-1]["brand_name"] == "Stone's Throw"


def test_parse_csv_dataset():
    csv_data = (
        "product id,brand name,class/type,abv,volume\n"
        "P-1,Old Tom,Bourbon,45% ABV,750 mL\n"
        ",No-ID Brand,Gin,40%,700 mL\n"
    ).encode()
    products = parse_dataset_file(csv_data, "catalog.csv")
    assert products[0]["product_id"] == "P-1"
    assert products[1]["product_id"].startswith("ROW-")  # auto-assigned
    assert products[1]["brand_name"] == "No-ID Brand"


def test_missing_brand_column_rejected():
    with pytest.raises(DatasetFileError, match="brand name"):
        parse_dataset_file(b"foo,bar\n1,2\n", "catalog.csv")


def test_wrong_extension_rejected():
    with pytest.raises(DatasetFileError, match="csv"):
        parse_dataset_file(b"whatever", "catalog.txt")


def test_expected_values_drops_id_and_blanks():
    values = expected_values(
        {"product_id": "P-1", "brand_name": "Old Tom", "class_type": "",
         "alcohol_content": "45%", "net_contents": "750 mL"}
    )
    assert "product_id" not in values
    assert "class_type" not in values
    assert values["brand_name"] == "Old Tom"
