"""Unit tests for the in-memory Excel export."""

import io

from openpyxl import load_workbook

from app.export import build_workbook
from app.schemas import CatalogMatch, FieldResult, LabelResult, WarningResult
from app.warning import GOVERNMENT_WARNING_TEXT


def _field(key, label, value, status="pass", notes=""):
    return FieldResult(
        field=key,
        label=label,
        value=value,
        present=value is not None,
        confidence="high",
        status=status,
        notes=notes,
    )


def _sample_results():
    passing = LabelResult(
        filename="good-label.png",
        verdict="pass",
        fields=[
            _field("brand_name", "Brand Name", "OLD TOM DISTILLERY"),
            _field("class_type", "Class / Type", "Kentucky Straight Bourbon Whiskey"),
            _field("alcohol_content", "Alcohol Content", "45% Alc./Vol."),
            _field("net_contents", "Net Contents", "750 mL"),
        ],
        warning_statement=WarningResult(
            status="pass",
            extracted_text=GOVERNMENT_WARNING_TEXT,
            expected_text=GOVERNMENT_WARNING_TEXT,
            notes=[],
        ),
        match=CatalogMatch(
            matched=True,
            method="embedding_retrieval",
            product={"product_id": "TTB-001", "brand_name": "Old Tom Distillery"},
            score=0.93,
        ),
        processing_seconds=3.21,
    )
    failing = LabelResult(
        filename="bad-label.png",
        verdict="fail",
        fields=[
            _field("brand_name", "Brand Name", "STONE'S THROW"),
            _field("class_type", "Class / Type", None, status="fail",
                   notes="Class / Type was not found on the label."),
            _field("alcohol_content", "Alcohol Content", "40% ABV"),
            _field("net_contents", "Net Contents", "750 mL"),
        ],
        warning_statement=WarningResult(
            status="fail",
            extracted_text="Government Warning: ...",
            expected_text=GOVERNMENT_WARNING_TEXT,
            notes=["The prefix must be in ALL CAPITAL letters."],
        ),
        match=CatalogMatch(
            matched=False,
            method="embedding_retrieval",
            score=0.41,
            note="No catalog product was similar enough.",
        ),
        processing_seconds=2.8,
    )
    errored = LabelResult(
        filename="cat.jpg",
        verdict="error",
        error="The image could not be read as an alcohol beverage label.",
    )
    return [passing, failing, errored]


_CATALOG = [
    {"product_id": "TTB-001", "brand_name": "Old Tom Distillery",
     "class_type": "Bourbon", "alcohol_content": "45% ABV", "net_contents": "750 mL"},
    {"product_id": "TTB-002", "brand_name": "Casamigos",
     "class_type": "Tequila", "alcohol_content": "40% ABV", "net_contents": "750 mL"},
]


def test_catalog_is_first_sheet_with_results_second():
    data = build_workbook(_sample_results(), _CATALOG)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Product Catalog", "Label Extraction Results"]
    catalog = wb["Product Catalog"]
    assert catalog["A1"].value == "Product ID"
    assert catalog["A2"].value == "TTB-001"
    assert catalog.max_row == 3  # header + 2 products


def test_matched_product_links_to_catalog_row():
    data = build_workbook(_sample_results(), _CATALOG)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Label Extraction Results"]
    # Row 2 matched TTB-001, which is catalog row 2.
    assert ws["I2"].hyperlink is not None
    assert ws["I2"].hyperlink.target == "#'Product Catalog'!A2"
    # Unmatched and errored rows get no link.
    assert ws["I3"].hyperlink is None
    assert ws["I4"].hyperlink is None


def test_workbook_layout_and_values():
    data = build_workbook(_sample_results(), _CATALOG)
    ws = load_workbook(io.BytesIO(data))["Label Extraction Results"]

    header = [cell.value for cell in ws[1]]
    assert header == [
        "Label File", "Brand Name", "Class / Type", "Alcohol Content",
        "Net Contents", "Bottler Name & Address", "Country of Origin",
        "Government Warning", "Matched Product", "Match Score",
        "Judge Confidence", "Verdict", "Notes", "Processing Time (s)",
    ]
    assert ws.max_row == 4  # header + one row per label

    assert ws["A2"].value == "good-label.png"
    assert ws["B2"].value == "OLD TOM DISTILLERY"
    # The warning column carries the verbatim transcription; a passing
    # warning gets NO fill — only problems are colored.
    assert ws["H2"].value == GOVERNMENT_WARNING_TEXT
    assert ws["H2"].fill.patternType is None
    assert ws["I2"].value == "TTB-001 Old Tom Distillery"
    assert ws["J2"].value == 0.93
    assert ws["L2"].value == "PASS"
    assert ws["N2"].value == 3.2

    assert ws["C3"].value == "NOT FOUND"
    assert ws["H3"].value == "Government Warning: ..."
    assert ws["H3"].fill.fgColor.rgb.endswith("FFC7CE")  # red: warning failed
    assert ws["I3"].value == "NO MATCH"
    assert ws["L3"].value == "FAIL"
    assert "ALL CAPITAL" in ws["M3"].value

    assert ws["L4"].value == "COULD NOT CHECK"
    assert "could not be read" in ws["M4"].value


def _tiny_png() -> bytes:
    import io as _io

    from PIL import Image

    buffer = _io.BytesIO()
    Image.new("RGB", (60, 40), "white").save(buffer, format="PNG")
    return buffer.getvalue()


def test_label_file_cells_link_to_embedded_images():
    images = {"good-label.png": _tiny_png(), "bad-label.png": _tiny_png()}
    data = build_workbook(_sample_results(), _CATALOG, images=images)
    wb = load_workbook(io.BytesIO(data))

    assert "Label Images" in wb.sheetnames
    images_ws = wb["Label Images"]
    assert images_ws["A2"].value == "good-label.png"
    assert len(images_ws._images) == 2  # two embedded thumbnails

    ws = wb["Label Extraction Results"]
    assert ws["A2"].hyperlink.target == "#'Label Images'!A2"
    assert ws["A3"].hyperlink.target == "#'Label Images'!A3"
    assert ws["A4"].hyperlink is None  # no image sent for cat.jpg


def test_no_images_means_no_images_sheet():
    data = build_workbook(_sample_results(), _CATALOG)
    wb = load_workbook(io.BytesIO(data))
    assert "Label Images" not in wb.sheetnames
    assert wb["Label Extraction Results"]["A2"].hyperlink is None


def test_batch_total_footer_row():
    data = build_workbook(_sample_results(), _CATALOG, total_seconds=12.34)
    ws = load_workbook(io.BytesIO(data))["Label Extraction Results"]
    footer = [cell.value for cell in ws[ws.max_row]]
    assert footer[-1] == 12.3
    assert "concurrently" in footer[-2]


def test_client_time_gets_its_own_footer_row():
    data = build_workbook(
        _sample_results(), _CATALOG, total_seconds=12.34, client_seconds=58.7
    )
    ws = load_workbook(io.BytesIO(data))["Label Extraction Results"]
    last = [cell.value for cell in ws[ws.max_row]]
    second_last = [cell.value for cell in ws[ws.max_row - 1]]
    assert second_last[-1] == 12.3 and "processing" in second_last[-2]
    assert last[-1] == 58.7 and "uploading" in last[-2]


def test_no_total_means_no_footer():
    data = build_workbook(_sample_results(), _CATALOG)
    ws = load_workbook(io.BytesIO(data))["Label Extraction Results"]
    assert ws.max_row == 4  # header + 3 labels, no footer


def test_no_catalog_means_results_only_workbook():
    data = build_workbook(_sample_results(), [])
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Label Extraction Results"]
    ws = wb.active
    assert ws["A2"].value == "good-label.png"
    # Match columns are dropped entirely when matching didn't run.
    header = [cell.value for cell in ws[1]]
    assert "Matched Product" not in header
    assert "Match Score" not in header
    assert "Judge Confidence" not in header
    assert header[-3:] == ["Verdict", "Notes", "Processing Time (s)"]


def test_verdict_cells_are_color_coded():
    data = build_workbook(_sample_results(), _CATALOG)
    ws = load_workbook(io.BytesIO(data))["Label Extraction Results"]

    assert ws["L2"].fill.fgColor.rgb.endswith("C6EFCE")  # green for pass
    assert ws["L3"].fill.fgColor.rgb.endswith("FFC7CE")  # red for fail
    assert ws["A1"].fill.fgColor.rgb.endswith("0B5394")  # styled header


def test_columns_have_explicit_widths():
    data = build_workbook(_sample_results(), _CATALOG)
    ws = load_workbook(io.BytesIO(data))["Label Extraction Results"]
    widths = [ws.column_dimensions[chr(ord("A") + i)].width for i in range(14)]
    assert all(w and w > 0 for w in widths)
